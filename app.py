import os, io, json, sqlite3, uuid, smtplib, ssl as ssl_lib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from functools import wraps
from flask import (Flask, request, jsonify, render_template,
                   redirect, url_for, session, flash, send_file)
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf.csrf import CSRFProtect, CSRFError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "change-me-2024")

# ─── SEGURANÇA DE SESSÃO E CSRF ───────────────────────────────────────────────
app.config["SESSION_COOKIE_HTTPONLY"]  = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"]   = os.getenv("HTTPS", "0") == "1"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
app.config["WTF_CSRF_TIME_LIMIT"] = 3600

csrf = CSRFProtect(app)

MAIL_SERVER = os.getenv("MAIL_SERVER", "")
MAIL_PORT   = int(os.getenv("MAIL_PORT", 465))
MAIL_USER   = os.getenv("MAIL_USERNAME", "")
MAIL_PASS   = os.getenv("MAIL_PASSWORD", "")
MAIL_SENDER = os.getenv("MAIL_DEFAULT_SENDER", MAIL_USER)

DB_PATH  = os.path.join(os.path.dirname(__file__), "app.db")
BASE_URL = os.getenv("BASE_URL", "http://localhost:5000")

# ─── RATE LIMIT DE LOGIN (em memória) ─────────────────────────────────────────
_login_attempts: dict = {}   # ip -> [datetime, ...]
_MAX_ATTEMPTS   = 5
_LOCKOUT_MIN    = 15

def _login_allowed(ip: str) -> bool:
    cutoff   = datetime.now() - timedelta(minutes=_LOCKOUT_MIN)
    attempts = [t for t in _login_attempts.get(ip, []) if t > cutoff]
    _login_attempts[ip] = attempts
    return len(attempts) < _MAX_ATTEMPTS

def _register_failed(ip: str):
    _login_attempts.setdefault(ip, []).append(datetime.now())

def _clear_attempts(ip: str):
    _login_attempts.pop(ip, None)

# ─── OCULTO_KEYS — todas as chaves que suportam toggle de visibilidade ────────
OCULTO_KEYS = (
    ["oculto_hero_badge", "oculto_hero_descricao",
     "oculto_secao_sobre", "oculto_secao_beneficios", "oculto_secao_cta",
     "oculto_footer_endereco", "oculto_footer_telefone"]
    + [f"oculto_stat{n}_num" for n in range(1, 4)]
    + [f"oculto_sobre_texto{n}" for n in range(1, 4)]
    + [f"oculto_tl{n}" for n in range(1, 5)]
    + [f"oculto_bene{n}" for n in range(1, 13)]
)

# ─── CONFIG PADRÃO ────────────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    "logo_sigla": "TC", "logo_nome": "TechConf SP",
    "logo_subtitulo": "Conferência de Tecnologia",
    "cor_primaria": "#0d2d6e", "cor_secundaria": "#1a4fad", "cor_destaque": "#00a86b",
    "meta_titulo": "TechConf SP 2024 — Inscrições Abertas",
    "meta_descricao": "O maior evento de tecnologia do Estado de São Paulo.",
    "hero_badge": "Inscrições Abertas · Edição 2024",
    "hero_titulo": "A conferência de tecnologia que vai transformar sua carreira",
    "hero_titulo_destaque": "tecnologia",
    "hero_descricao": "Três dias de imersão com os maiores especialistas do setor. Palestras, workshops, networking e muito mais em São Paulo.",
    "stat1_num": "500+", "stat1_label": "Participantes",
    "stat2_num": "3",    "stat2_label": "Dias de evento",
    "stat3_num": "40+",  "stat3_label": "Palestrantes",
    "form_titulo": "Garanta sua vaga",
    "form_subtitulo": "Preencha o formulário e receba a confirmação no seu e-mail.",
    "form_botao": "Quero me inscrever",
    "sobre_titulo": "Sobre o Evento", "sobre_subtitulo": "Uma história de inovação e conexão",
    "sobre_texto1": "A TechConf SP nasceu em 2018 da vontade de um grupo de profissionais de tecnologia de criar um espaço genuíno de troca de conhecimento e experiências no Estado de São Paulo.",
    "sobre_texto2": "Ao longo das edições, o evento cresceu e se tornou referência no setor, reunindo profissionais de todo o Brasil para debater as tendências e desafios do mercado de tecnologia.",
    "sobre_texto3": "Em 2024, chegamos à nossa maior edição, com programação inédita, palestrantes internacionais e um hub de inovação dedicado ao futuro da tecnologia.",
    "tl1_icone": "⚡", "tl1_titulo": "2018 — Primeira edição",
    "tl1_texto": "Evento piloto com 80 participantes e 10 palestrantes locais.",
    "tl2_icone": "📈", "tl2_titulo": "2019–2021 — Crescimento",
    "tl2_texto": "Três edições consecutivas com crescimento médio de 60% no número de participantes.",
    "tl3_icone": "🏆", "tl3_titulo": "2022 — Reconhecimento nacional",
    "tl3_texto": "Eleita um dos 10 melhores eventos de tecnologia do Brasil.",
    "tl4_icone": "🚀", "tl4_titulo": "2024 — Nova era",
    "tl4_texto": "Nossa maior edição: 500+ participantes e palestrantes internacionais.",
    "bene_titulo": "Por que participar?",
    "bene_subtitulo": "Tudo que você vai encontrar na TechConf SP 2024",
    "bene1_icone": "🎤", "bene1_titulo": "Palestras Exclusivas",
    "bene1_texto": "Conteúdo de alto nível com especialistas reconhecidos no mercado nacional e internacional.",
    "bene2_icone": "🤝", "bene2_titulo": "Networking Qualificado",
    "bene2_texto": "Conecte-se com profissionais, recrutadores e líderes das maiores empresas do Brasil.",
    "bene3_icone": "🎓", "bene3_titulo": "Workshops Práticos",
    "bene3_texto": "Sessões hands-on para aprender e aplicar novas tecnologias com especialistas.",
    "bene4_icone": "💡", "bene4_titulo": "Inovação & Tendências",
    "bene4_texto": "Fique por dentro das tecnologias que vão moldar o futuro: IA, Cloud, Blockchain e mais.",
    "bene5_icone": "🏅", "bene5_titulo": "Certificado de Participação",
    "bene5_texto": "Certificado reconhecido pelo mercado para agregar valor ao seu currículo.",
    "bene6_icone": "🎁", "bene6_titulo": "Kit Exclusivo",
    "bene6_texto": "Todos os participantes recebem um kit especial com materiais e brindes.",
    # Benefícios 7–12 (vazios por padrão, cliente adiciona via "+")
    **{f"bene{n}_{k}": "" for n in range(7, 13) for k in ["icone", "titulo", "texto"]},
    "cta_titulo": "Não fique de fora!",
    "cta_subtitulo": "As vagas são limitadas. Garanta a sua inscrição agora mesmo.",
    "cta_botao": "Inscreva-se gratuitamente",
    "footer_texto": "TechConf SP · Conferência de Tecnologia · São Paulo, 2024",
    "footer_endereco": "",
    "footer_telefone": "",
    "email_assunto": "Confirme sua inscrição – TechConf SP 2024",
    "email_remetente_nome": "TechConf SP",
    "email_template_confirmacao_id": "",
}

# ─── TEMPLATES DE E-MAIL PADRÃO ───────────────────────────────────────────────
def _tpl(cor, conteudo_html):
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f6fb;font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6fb;padding:40px 16px;">
<tr><td align="center">
<table width="100%" cellpadding="0" cellspacing="0"
       style="max-width:560px;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.1);">
  <tr><td style="background:{cor};padding:32px 40px;text-align:center;">
    <h1 style="margin:0;color:#fff;font-size:1.3rem;font-weight:800;">{{{{evento}}}}</h1>
  </td></tr>
  <tr><td style="padding:36px 40px 28px;">
    <p style="font-size:1rem;color:#1e2637;font-weight:700;margin:0 0 12px;">Olá, {{{{nome}}}}!</p>
    {conteudo_html}
  </td></tr>
  <tr><td style="background:#f4f6fb;padding:18px 40px;text-align:center;border-top:1px solid #e5e7eb;">
    <p style="font-size:.78rem;color:#9ca3af;margin:0;">
      Você recebe este e-mail por estar cadastrado em {{{{evento}}}}.
    </p>
  </td></tr>
</table></td></tr></table></body></html>"""

EMAIL_TEMPLATES_PADRAO = [
    (
        "Confirme sua inscrição",
        "Confirme sua inscrição – {{evento}}",
        _tpl("#0d2d6e",
             '<p style="font-size:.95rem;color:#6b7280;line-height:1.7;margin:0 0 24px;">'
             'Para ativar sua inscrição, clique no botão abaixo:</p>'
             '<div style="text-align:center;margin:28px 0;">'
             '<a href="{{link}}" style="background:#00a86b;color:#fff;text-decoration:none;'
             'padding:14px 36px;border-radius:99px;font-size:1rem;font-weight:700;display:inline-block;">'
             'Confirmar minha inscrição</a></div>'
             '<p style="font-size:.8rem;color:#9ca3af;margin:0;">Ou copie: {{link}}</p>'),
    ),
    (
        "Sua inscrição foi realizada",
        "Inscrição recebida com sucesso – {{evento}}",
        _tpl("#1a4fad",
             '<p style="font-size:.95rem;color:#6b7280;line-height:1.7;margin:0;">'
             'Sua inscrição foi <strong style="color:#065f46;">recebida com sucesso</strong>! '
             'Estamos felizes em tê-lo(a) conosco.<br><br>'
             'Em breve você receberá mais informações sobre o evento. Fique de olho no seu e-mail.</p>'),
    ),
    (
        "Inscrição cancelada",
        "Sua inscrição foi cancelada – {{evento}}",
        _tpl("#6b7280",
             '<p style="font-size:.95rem;color:#6b7280;line-height:1.7;margin:0;">'
             'Informamos que sua inscrição em <strong>{{evento}}</strong> foi <strong style="color:#dc2626;">cancelada</strong>.<br><br>'
             'Se isso foi um engano ou deseja se reinscrever, entre em contato conosco.</p>'),
    ),
]

EMAIL_TEMPLATE_NOVO = _tpl(
    "#0d2d6e",
    '<p style="font-size:.95rem;color:#6b7280;line-height:1.7;margin:0;">'
    'Escreva aqui o conteúdo do e-mail. Use <strong>{{nome}}</strong> para o nome do destinatário.</p>'
)


# ─── DATABASE ─────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS inscricoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE, telefone TEXT NOT NULL,
            token TEXT NOT NULL UNIQUE, confirmado INTEGER DEFAULT 0,
            cancelado INTEGER DEFAULT 0, na_lixeira INTEGER DEFAULT 0,
            deletado_em TEXT, dados_extras TEXT, criado_em TEXT NOT NULL)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS configuracoes (
            chave TEXT PRIMARY KEY, valor TEXT NOT NULL)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS campos_formulario (
            id INTEGER PRIMARY KEY AUTOINCREMENT, label TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'text', placeholder TEXT DEFAULT '',
            obrigatorio INTEGER DEFAULT 0, ordem INTEGER DEFAULT 0)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS email_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT NOT NULL,
            assunto TEXT NOT NULL, corpo TEXT NOT NULL,
            criado_em TEXT NOT NULL, atualizado_em TEXT NOT NULL)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS email_campanhas (
            id INTEGER PRIMARY KEY AUTOINCREMENT, template_id INTEGER NOT NULL,
            template_nome TEXT NOT NULL, assunto TEXT NOT NULL,
            grupo TEXT NOT NULL, total_enviados INTEGER DEFAULT 0,
            enviado_em TEXT NOT NULL)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            criado_em TEXT NOT NULL)""")
        # Migrações seguras
        for col, dfn in [("cancelado","INTEGER DEFAULT 0"),("na_lixeira","INTEGER DEFAULT 0"),
                         ("deletado_em","TEXT"),("dados_extras","TEXT"),("numero_inscricao","TEXT")]:
            try: conn.execute(f"ALTER TABLE inscricoes ADD COLUMN {col} {dfn}")
            except: pass
        # Semear templates padrão
        if conn.execute("SELECT COUNT(*) FROM email_templates").fetchone()[0] == 0:
            agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for nome_t, assunto_t, corpo_t in EMAIL_TEMPLATES_PADRAO:
                conn.execute("INSERT INTO email_templates (nome,assunto,corpo,criado_em,atualizado_em)"
                             " VALUES (?,?,?,?,?)", (nome_t, assunto_t, corpo_t, agora, agora))
        conn.commit()


def init_admin():
    """Cria o admin inicial a partir do .env — só executa se nenhum admin existir."""
    email = os.getenv("ADMIN_EMAIL", "").strip().lower()
    senha = os.getenv("ADMIN_PASSWORD", "").strip()
    if not email or not senha:
        return
    with get_db() as conn:
        if conn.execute("SELECT COUNT(*) FROM admins").fetchone()[0] == 0:
            conn.execute(
                "INSERT INTO admins (email, password_hash, criado_em) VALUES (?,?,?)",
                (email, generate_password_hash(senha),
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()


def get_config():
    cfg = dict(DEFAULT_CONFIG)
    with get_db() as conn:
        for row in conn.execute("SELECT chave, valor FROM configuracoes").fetchall():
            cfg[row["chave"]] = row["valor"]
    return cfg


def save_config(data: dict):
    with get_db() as conn:
        for k, v in data.items():
            conn.execute("INSERT INTO configuracoes (chave,valor) VALUES (?,?) "
                         "ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor", (k, v))
        conn.commit()


def get_campos_ativos():
    with get_db() as conn:
        return conn.execute(
            "SELECT * FROM campos_formulario ORDER BY ordem, id").fetchall()


# ─── HEADERS DE SEGURANÇA ────────────────────────────────────────────────────

@app.after_request
def add_security_headers(response):
    response.headers["X-Frame-Options"]        = "SAMEORIGIN"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-XSS-Protection"]       = "1; mode=block"
    response.headers["Referrer-Policy"]        = "strict-origin-when-cross-origin"
    return response


@app.errorhandler(CSRFError)
def csrf_error(e):
    flash("Token de segurança inválido. Recarregue a página e tente novamente.", "erro")
    return redirect(request.referrer or url_for("admin_painel")), 400


# ─── CONTEXT PROCESSOR ───────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    counts = {"lixeira_count": 0}
    if session.get("admin_logged_in"):
        with get_db() as conn:
            counts["lixeira_count"] = conn.execute(
                "SELECT COUNT(*) FROM inscricoes WHERE na_lixeira=1").fetchone()[0]
    return counts


# ─── EMAIL ────────────────────────────────────────────────────────────────────

def send_email(to_addr, subject, html_body):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = MAIL_SENDER
    msg["To"]      = to_addr
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    ctx = ssl_lib.create_default_context()
    with smtplib.SMTP_SSL(MAIL_SERVER, MAIL_PORT, context=ctx, timeout=15) as s:
        s.login(MAIL_USER, MAIL_PASS)
        s.sendmail(MAIL_USER, to_addr, msg.as_string())


def render_template_vars(body: str, lead: dict, extra: dict = None) -> str:
    result = (body.replace("{{nome}}", lead.get("nome",""))
                  .replace("{{email}}", lead.get("email","")))
    if extra:
        for k, v in extra.items():
            result = result.replace(f"{{{{{k}}}}}", v)
    return result


def _build_confirmacao_email(cfg, nome, link):
    """Retorna (assunto, html_body) do e-mail de confirmação."""
    tpl_id = cfg.get("email_template_confirmacao_id", "")
    if tpl_id:
        with get_db() as conn:
            t = conn.execute("SELECT * FROM email_templates WHERE id=?",
                             (int(tpl_id),)).fetchone()
        if t:
            body = render_template_vars(t["corpo"], {"nome": nome},
                                        {"link": link, "evento": cfg.get("logo_nome","Evento")})
            return t["assunto"].replace("{{evento}}", cfg.get("logo_nome","Evento")), body
    # Fallback: template HTML estático
    html = render_template("email_confirmar.html", nome=nome, link=link,
        site_nome=cfg.get("logo_nome","Evento"), site_sigla=cfg.get("logo_sigla","EV"),
        site_subtitulo=cfg.get("logo_subtitulo",""),
        cor_primaria=cfg.get("cor_primaria","#0d2d6e"),
        cor_destaque=cfg.get("cor_destaque","#00a86b"))
    return cfg.get("email_assunto","Confirme sua inscrição"), html


# ─── AUTH ─────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


# ─── PUBLIC ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    cfg = get_config()
    titulo   = cfg.get("hero_titulo","")
    destaque = cfg.get("hero_titulo_destaque","")
    cfg["hero_titulo_html"] = (
        titulo.replace(destaque, f"<span>{destaque}</span>", 1)
        if destaque and destaque in titulo else titulo)
    campos = get_campos_ativos()
    return render_template("index.html", c=cfg, campos=campos)


@app.route("/inscrever", methods=["POST"])
@csrf.exempt
def inscrever():
    data     = request.get_json()
    nome     = (data.get("nome") or "").strip()
    email    = (data.get("email") or "").strip().lower()
    telefone = (data.get("telefone") or "").strip()
    if not nome or not email or not telefone:
        return jsonify({"ok": False, "erro": "Preencha todos os campos."}), 400

    # Valida campos extras obrigatórios
    campos = get_campos_ativos()
    extras_raw = data.get("extras", {})
    for c in campos:
        val = (extras_raw.get(str(c["id"])) or "").strip()
        if c["obrigatorio"] and not val:
            return jsonify({"ok": False, "erro": f'O campo "{c["label"]}" é obrigatório.'}), 400
    dados_extras = json.dumps(extras_raw) if extras_raw else None

    token     = str(uuid.uuid4())
    criado_em = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO inscricoes (nome,email,telefone,token,dados_extras,criado_em)"
                " VALUES (?,?,?,?,?,?)",
                (nome, email, telefone, token, dados_extras, criado_em))
            conn.commit()
    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "erro": "Este e-mail já está cadastrado."}), 409

    cfg  = get_config()
    link = f"{BASE_URL}/confirmar/{token}"
    try:
        assunto, html_body = _build_confirmacao_email(cfg, nome, link)
        send_email(email, assunto, html_body)
    except Exception as e:
        app.logger.error(f"E-mail erro: {e}")
    return jsonify({"ok": True})


@app.route("/confirmar/<token>")
def confirmar(token):
    import random
    cfg = get_config()
    with get_db() as conn:
        row = conn.execute("SELECT * FROM inscricoes WHERE token=?", (token,)).fetchone()
        if not row:
            return render_template("confirmar.html", status="invalido", c=cfg)
        if row["confirmado"]:
            return render_template("confirmar.html", status="ja_confirmado", nome=row["nome"],
                                   numero=row["numero_inscricao"], c=cfg)
        numero = f"#{random.randint(1000, 9999)}"
        conn.execute("UPDATE inscricoes SET confirmado=1, numero_inscricao=? WHERE token=?",
                     (numero, token))
        conn.commit()
    try:
        evento = cfg.get("logo_nome", "Evento")
        assunto = f"Inscrição confirmada – {evento}"
        html = f"""
        <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:560px;margin:0 auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.1);">
          <div style="background:{cfg.get('cor_primaria','#0d2d6e')};padding:32px 40px;text-align:center;">
            <h1 style="margin:0;color:#fff;font-size:1.3rem;font-weight:800;">{evento}</h1>
          </div>
          <div style="padding:36px 40px;">
            <p style="font-size:1rem;color:#1e2637;font-weight:700;margin:0 0 12px;">Olá, {row['nome']}!</p>
            <p style="font-size:.95rem;color:#6b7280;line-height:1.7;margin:0 0 24px;">
              Sua inscrição foi <strong style="color:#065f46;">confirmada com sucesso</strong>!
              Estamos felizes em tê-lo(a) conosco.
            </p>
            <div style="background:#f4f6fb;border-radius:12px;padding:20px;text-align:center;margin:24px 0;">
              <p style="margin:0;font-size:.85rem;color:#6b7280;">Seu número de inscrição</p>
              <p style="margin:8px 0 0;font-size:2rem;font-weight:900;color:{cfg.get('cor_primaria','#0d2d6e')};letter-spacing:2px;">{numero}</p>
            </div>
            <p style="font-size:.85rem;color:#9ca3af;margin:0;">
              Guarde este número — ele identifica sua inscrição em {evento}.
            </p>
          </div>
          <div style="background:#f4f6fb;padding:18px 40px;text-align:center;border-top:1px solid #e5e7eb;">
            <p style="font-size:.78rem;color:#9ca3af;margin:0;">
              Você recebe este e-mail por estar inscrito em {evento}.
            </p>
          </div>
        </div>"""
        send_email(row["email"], assunto, html)
    except Exception as e:
        app.logger.error(f"E-mail confirmação erro: {e}")
    return render_template("confirmar.html", status="ok", nome=row["nome"], numero=numero, c=cfg)


# ─── ADMIN AUTH ───────────────────────────────────────────────────────────────

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    erro = None
    if request.method == "POST":
        ip = request.remote_addr
        if not _login_allowed(ip):
            erro = f"Muitas tentativas. Aguarde {_LOCKOUT_MIN} minutos."
            return render_template("admin_login.html", erro=erro)

        email_input = request.form.get("email", "").strip().lower()
        senha_input = request.form.get("senha", "")

        with get_db() as conn:
            admin = conn.execute(
                "SELECT * FROM admins WHERE email=?", (email_input,)).fetchone()

        if admin and check_password_hash(admin["password_hash"], senha_input):
            _clear_attempts(ip)
            session["admin_logged_in"] = True
            session["admin_email"]     = admin["email"]
            session.permanent          = True
            return redirect(url_for("admin_painel"))

        _register_failed(ip)
        erro = "E-mail ou senha incorretos."
    return render_template("admin_login.html", erro=erro)


@app.route("/admin/alterar-senha", methods=["GET", "POST"])
@login_required
def admin_alterar_senha():
    erro = None
    if request.method == "POST":
        atual    = request.form.get("atual", "")
        nova     = request.form.get("nova", "")
        confirma = request.form.get("confirma", "")
        email    = session.get("admin_email", "")
        with get_db() as conn:
            admin = conn.execute(
                "SELECT * FROM admins WHERE email=?", (email,)).fetchone()
        if not admin or not check_password_hash(admin["password_hash"], atual):
            erro = "Senha atual incorreta."
        elif len(nova) < 8:
            erro = "A nova senha deve ter pelo menos 8 caracteres."
        elif nova != confirma:
            erro = "As senhas não coincidem."
        else:
            with get_db() as conn:
                conn.execute("UPDATE admins SET password_hash=? WHERE email=?",
                             (generate_password_hash(nova), email))
                conn.commit()
            flash("Senha alterada com sucesso!", "ok")
            return redirect(url_for("admin_painel"))
    cfg = get_config()
    return render_template("admin_senha.html", erro=erro, c=cfg)


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


# ─── ADMIN LEADS ─────────────────────────────────────────────────────────────

@app.route("/admin")
@login_required
def admin_painel():
    filtro = request.args.get("filtro","todos")
    cfg    = get_config()
    base   = "WHERE na_lixeira=0"
    q = {"confirmados": base+" AND confirmado=1 AND cancelado=0",
         "pendentes":   base+" AND confirmado=0 AND cancelado=0",
         "cancelados":  base+" AND cancelado=1"}.get(filtro, base)
    with get_db() as conn:
        rows        = conn.execute(f"SELECT * FROM inscricoes {q} ORDER BY criado_em DESC").fetchall()
        total       = conn.execute("SELECT COUNT(*) FROM inscricoes WHERE na_lixeira=0").fetchone()[0]
        confirmados = conn.execute("SELECT COUNT(*) FROM inscricoes WHERE confirmado=1 AND cancelado=0 AND na_lixeira=0").fetchone()[0]
        pendentes   = conn.execute("SELECT COUNT(*) FROM inscricoes WHERE confirmado=0 AND cancelado=0 AND na_lixeira=0").fetchone()[0]
        cancelados  = conn.execute("SELECT COUNT(*) FROM inscricoes WHERE cancelado=1 AND na_lixeira=0").fetchone()[0]
    campos = get_campos_ativos()
    return render_template("admin_leads.html", inscricoes=rows, total=total,
        confirmados=confirmados, pendentes=pendentes, cancelados=cancelados,
        filtro=filtro, c=cfg, campos=campos)


@app.route("/admin/mover-lixeira/<int:id>", methods=["POST"])
@login_required
def admin_mover_lixeira(id):
    with get_db() as conn:
        conn.execute("UPDATE inscricoes SET na_lixeira=1, deletado_em=? WHERE id=?",
                     (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), id))
        conn.commit()
    flash("Lead movido para a lixeira.", "ok")
    return redirect(request.referrer or url_for("admin_painel"))


@app.route("/admin/cancelar/<int:id>", methods=["POST"])
@login_required
def admin_cancelar(id):
    with get_db() as conn:
        row = conn.execute("SELECT cancelado FROM inscricoes WHERE id=?", (id,)).fetchone()
        if row:
            novo = 0 if row["cancelado"] else 1
            conn.execute("UPDATE inscricoes SET cancelado=? WHERE id=?", (novo, id))
            conn.commit()
            flash("Inscrição cancelada." if novo else "Inscrição reativada.", "ok")
    return redirect(request.referrer or url_for("admin_painel"))


@app.route("/admin/exportar")
@login_required
def admin_exportar():
    filtro = request.args.get("filtro","todos")
    base   = "WHERE na_lixeira=0"
    q = {"confirmados": base+" AND confirmado=1 AND cancelado=0",
         "pendentes":   base+" AND confirmado=0 AND cancelado=0",
         "cancelados":  base+" AND cancelado=1"}.get(filtro, base)
    campos = get_campos_ativos()
    with get_db() as conn:
        rows = conn.execute(f"SELECT * FROM inscricoes {q} ORDER BY criado_em DESC").fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    hdr_fill = PatternFill("solid", fgColor="0D2D6E")
    hdr_font = Font(color="FFFFFF", bold=True)
    headers = ["ID","Nº Inscrição","Nome","E-mail","Telefone","Confirmado","Cancelado","Cadastro"]
    headers += [c["label"] for c in campos]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        extras = json.loads(row["dados_extras"] or "{}") if row["dados_extras"] else {}
        ws.cell(r,1,row["id"]); ws.cell(r,2,row["numero_inscricao"] or "")
        ws.cell(r,3,row["nome"]); ws.cell(r,4,row["email"])
        ws.cell(r,5,row["telefone"])
        ws.cell(r,6,"Sim" if row["confirmado"] else "Não")
        ws.cell(r,7,"Sim" if row["cancelado"]  else "Não")
        ws.cell(r,8,row["criado_em"])
        for col, c in enumerate(campos, 9):
            ws.cell(r, col, extras.get(str(c["id"]),""))
    for col, w in enumerate([8,16,30,34,18,12,12,22]+[20]*len(campos), 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"leads_{filtro}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── ADMIN LIXEIRA ────────────────────────────────────────────────────────────

@app.route("/admin/lixeira")
@login_required
def admin_lixeira():
    cfg = get_config()
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM inscricoes WHERE na_lixeira=1 ORDER BY deletado_em DESC").fetchall()
    return render_template("admin_lixeira.html", inscricoes=rows, c=cfg)


@app.route("/admin/restaurar/<int:id>", methods=["POST"])
@login_required
def admin_restaurar(id):
    with get_db() as conn:
        conn.execute("UPDATE inscricoes SET na_lixeira=0, deletado_em=NULL WHERE id=?", (id,))
        conn.commit()
    flash("Lead restaurado.", "ok")
    return redirect(url_for("admin_lixeira"))


@app.route("/admin/deletar-permanente/<int:id>", methods=["POST"])
@login_required
def admin_deletar_permanente(id):
    with get_db() as conn:
        conn.execute("DELETE FROM inscricoes WHERE id=? AND na_lixeira=1", (id,))
        conn.commit()
    flash("Lead excluído permanentemente.", "ok")
    return redirect(url_for("admin_lixeira"))


@app.route("/admin/esvaziar-lixeira", methods=["POST"])
@login_required
def admin_esvaziar_lixeira():
    with get_db() as conn:
        conn.execute("DELETE FROM inscricoes WHERE na_lixeira=1")
        conn.commit()
    flash("Lixeira esvaziada.", "ok")
    return redirect(url_for("admin_lixeira"))


# ─── ADMIN CAMPOS DO FORMULÁRIO ───────────────────────────────────────────────

@app.route("/admin/campos/adicionar", methods=["POST"])
@login_required
def admin_campo_adicionar():
    label       = request.form.get("label","").strip()
    tipo        = request.form.get("tipo","text")
    placeholder = request.form.get("placeholder","").strip()
    obrigatorio = 1 if request.form.get("obrigatorio") else 0
    if label:
        with get_db() as conn:
            ordem = (conn.execute("SELECT COUNT(*) FROM campos_formulario").fetchone()[0])
            conn.execute("INSERT INTO campos_formulario (label,tipo,placeholder,obrigatorio,ordem)"
                         " VALUES (?,?,?,?,?)", (label, tipo, placeholder, obrigatorio, ordem))
            conn.commit()
        flash(f'Campo "{label}" adicionado.', "ok")
    return redirect(url_for("admin_personalizar") + "#sec-campos")


@app.route("/admin/campos/<int:id>/deletar", methods=["POST"])
@login_required
def admin_campo_deletar(id):
    with get_db() as conn:
        conn.execute("DELETE FROM campos_formulario WHERE id=?", (id,))
        conn.commit()
    flash("Campo removido.", "ok")
    return redirect(url_for("admin_personalizar") + "#sec-campos")


# ─── ADMIN EMAILS ─────────────────────────────────────────────────────────────

@app.route("/admin/emails")
@login_required
def admin_emails():
    cfg = get_config()
    with get_db() as conn:
        templates = conn.execute("SELECT * FROM email_templates ORDER BY atualizado_em DESC").fetchall()
        campanhas = conn.execute("SELECT * FROM email_campanhas ORDER BY enviado_em DESC LIMIT 20").fetchall()
    return render_template("admin_emails.html", templates=templates, campanhas=campanhas, c=cfg)


@app.route("/admin/emails/novo", methods=["GET","POST"])
@login_required
def admin_email_novo():
    cfg = get_config()
    if request.method == "POST":
        nome    = request.form.get("nome","").strip()
        assunto = request.form.get("assunto","").strip()
        corpo   = request.form.get("corpo","").strip()
        if not nome or not assunto or not corpo:
            flash("Preencha todos os campos.", "erro")
            return render_template("admin_email_editor.html", c=cfg, t=None,
                                   form=request.form)
        agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as conn:
            conn.execute("INSERT INTO email_templates (nome,assunto,corpo,criado_em,atualizado_em)"
                         " VALUES (?,?,?,?,?)", (nome,assunto,corpo,agora,agora))
            conn.commit()
        flash(f'Template "{nome}" criado!', "ok")
        return redirect(url_for("admin_emails"))
    return render_template("admin_email_editor.html", c=cfg, t=None,
                           form={"nome":"","assunto":"","corpo":EMAIL_TEMPLATE_NOVO})


@app.route("/admin/emails/<int:id>/editar", methods=["GET","POST"])
@login_required
def admin_email_editar(id):
    cfg = get_config()
    with get_db() as conn:
        t = conn.execute("SELECT * FROM email_templates WHERE id=?", (id,)).fetchone()
    if not t:
        flash("Template não encontrado.", "erro")
        return redirect(url_for("admin_emails"))
    if request.method == "POST":
        agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as conn:
            conn.execute("UPDATE email_templates SET nome=?,assunto=?,corpo=?,atualizado_em=? WHERE id=?",
                         (request.form.get("nome",""), request.form.get("assunto",""),
                          request.form.get("corpo",""), agora, id))
            conn.commit()
        flash("Template atualizado.", "ok")
        return redirect(url_for("admin_emails"))
    return render_template("admin_email_editor.html", c=cfg, t=t, form=t)


@app.route("/admin/emails/<int:id>/deletar", methods=["POST"])
@login_required
def admin_email_deletar(id):
    with get_db() as conn:
        conn.execute("DELETE FROM email_templates WHERE id=?", (id,))
        conn.commit()
    flash("Template excluído.", "ok")
    return redirect(url_for("admin_emails"))


@app.route("/admin/emails/<int:id>/enviar", methods=["GET","POST"])
@login_required
def admin_email_enviar(id):
    cfg = get_config()
    with get_db() as conn:
        t = conn.execute("SELECT * FROM email_templates WHERE id=?", (id,)).fetchone()
    if not t:
        flash("Template não encontrado.", "erro")
        return redirect(url_for("admin_emails"))
    GRUPOS = {
        "confirmados": ("Confirmados",    "WHERE confirmado=1 AND cancelado=0 AND na_lixeira=0"),
        "pendentes":   ("Pendentes",      "WHERE confirmado=0 AND cancelado=0 AND na_lixeira=0"),
        "ativos":      ("Todos os ativos","WHERE cancelado=0 AND na_lixeira=0"),
        "cancelados":  ("Cancelados",     "WHERE cancelado=1 AND na_lixeira=0"),
    }
    if request.method == "POST":
        grupo = request.form.get("grupo","ativos")
        if grupo not in GRUPOS: flash("Grupo inválido.","erro"); return redirect(url_for("admin_email_enviar",id=id))
        _, filtro_sql = GRUPOS[grupo]
        with get_db() as conn:
            leads = conn.execute(f"SELECT nome,email FROM inscricoes {filtro_sql}").fetchall()
        enviados = erros = 0
        for lead in leads:
            corpo = render_template_vars(t["corpo"], dict(lead),
                                        {"evento": cfg.get("logo_nome","Evento")})
            try: send_email(lead["email"], t["assunto"], corpo); enviados += 1
            except Exception as e: app.logger.error(f"Erro {lead['email']}: {e}"); erros += 1
        agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as conn:
            conn.execute("INSERT INTO email_campanhas (template_id,template_nome,assunto,grupo,total_enviados,enviado_em)"
                         " VALUES (?,?,?,?,?,?)",
                         (id, t["nome"], t["assunto"], GRUPOS[grupo][0], enviados, agora))
            conn.commit()
        flash(f"{enviados} e-mails enviados" + (f", {erros} falhas." if erros else "."), "ok")
        return redirect(url_for("admin_emails"))
    with get_db() as conn:
        contagens = {k: (lbl, conn.execute(f"SELECT COUNT(*) FROM inscricoes {sql}").fetchone()[0])
                     for k,(lbl,sql) in GRUPOS.items()}
    return render_template("admin_email_enviar.html", c=cfg, t=t, grupos=GRUPOS, contagens=contagens)


# ─── ADMIN PERSONALIZAR ───────────────────────────────────────────────────────

@app.route("/admin/personalizar", methods=["GET","POST"])
@login_required
def admin_personalizar():
    cfg    = get_config()
    campos = get_campos_ativos()
    with get_db() as conn:
        email_templates = conn.execute("SELECT id,nome FROM email_templates ORDER BY nome").fetchall()
    if request.method == "POST":
        allowed = set(DEFAULT_CONFIG.keys()) | {"email_template_confirmacao_id"}
        data = {k: v for k, v in request.form.items() if k in allowed}
        # Salva oculto_* explicitamente (checkbox desmarcado não vem no form)
        for key in OCULTO_KEYS:
            data[key] = "1" if request.form.get(key) == "1" else "0"
        save_config(data)
        flash("Página atualizada!", "ok")
        return redirect(url_for("admin_personalizar"))
    return render_template("admin_personalizar.html", c=cfg, campos=campos,
                           email_templates=email_templates)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

# Inicializa banco e admin ao importar o módulo (gunicorn, Railway, Passenger)
init_db()
init_admin()

if __name__ == "__main__":
    app.run(debug=os.getenv("FLASK_DEBUG", "0") == "1")
